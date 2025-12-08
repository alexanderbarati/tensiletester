#!/usr/bin/env python3
"""
Export System for Professional Tensile Testing System

Comprehensive export to CSV, Excel, JSON, PDF, and XML formats
with full test data, metadata, and mechanical properties.

Author: DIY Tensile Tester Project
Version: 2.0.0
"""

import os
import json
from datetime import datetime
from typing import List, Optional, Dict, Any
from dataclasses import asdict
import numpy as np

from models import (
    TestConfiguration, MechanicalProperties, TestResults,
    TestMetadata, SpecimenConfig, ExportConfig
)

# Try to import optional dependencies
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ExportError(Exception):
    """Export operation error."""
    pass


class DataExporter:
    """Handles export of test data to various formats."""
    
    def __init__(self, output_dir: str = "./results"):
        self.output_dir = output_dir
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Ensure output directory exists."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def generate_filename(self, config: TestConfiguration, extension: str) -> str:
        """Generate filename from pattern: date-material_name."""
        timestamp = datetime.now()
        
        # Get material name, fallback to sample_id, then "unknown"
        material = config.metadata.material_name or config.metadata.sample_id or "unknown"
        
        # Clean material name for filename
        material = "".join(c if c.isalnum() or c in "-_" else "_" for c in material)
        
        # Format: YYYYMMDD-material_name.ext
        filename = f"{timestamp.strftime('%Y%m%d')}-{material}.{extension}"
        return os.path.join(self.output_dir, filename)
    
    def generate_default_filename(self, config: TestConfiguration, extension: str) -> str:
        """Generate default filename suggestion."""
        timestamp = datetime.now()
        material = config.metadata.material_name or config.metadata.sample_id or "unknown"
        material = "".join(c if c.isalnum() or c in "-_" else "_" for c in material)
        return f"{timestamp.strftime('%Y%m%d')}-{material}.{extension}"
    
    # ============== CSV Export ==============
    
    def export_csv(self, 
                   times: List[float],
                   forces: List[float],
                   extensions: List[float],
                   stresses: List[float],
                   strains: List[float],
                   config: TestConfiguration,
                   properties: MechanicalProperties,
                   include_header: bool = True,
                   filepath: str = None) -> str:
        """Export test data to CSV file."""
        filename = filepath if filepath else self.generate_filename(config, "csv")
        
        try:
            with open(filename, 'w', newline='') as f:
                if include_header:
                    # Write metadata header
                    f.write(f"# Tensile Test Results\n")
                    f.write(f"# Sample ID: {config.metadata.sample_id}\n")
                    f.write(f"# Material: {config.metadata.material_name}\n")
                    f.write(f"# Standard: {config.metadata.test_standard.value}\n")
                    f.write(f"# Date: {config.metadata.test_date}\n")
                    f.write(f"# Operator: {config.metadata.operator_name}\n")
                    f.write(f"# Gauge Length: {config.specimen.gauge_length} mm\n")
                    f.write(f"# Cross-Section: {config.specimen.cross_section_area} mm²\n")
                    f.write(f"#\n")
                    f.write(f"# Results:\n")
                    f.write(f"# UTS: {properties.ultimate_tensile_strength:.2f} MPa\n")
                    f.write(f"# Yield Strength: {properties.yield_strength_offset:.2f} MPa\n")
                    f.write(f"# Young's Modulus: {properties.youngs_modulus:.1f} MPa\n")
                    f.write(f"# Elongation: {properties.strain_at_break:.2f} %\n")
                    f.write(f"#\n")
                
                # Column headers
                f.write("Time (s),Force (N),Extension (mm),Stress (MPa),Strain (%)\n")
                
                # Data rows
                for i in range(len(times)):
                    f.write(f"{times[i]:.4f},{forces[i]:.4f},{extensions[i]:.4f},"
                           f"{stresses[i]:.4f},{strains[i]*100:.4f}\n")
            
            return filename
            
        except Exception as e:
            raise ExportError(f"CSV export failed: {e}")
    
    # ============== Excel Export ==============
    
    def export_excel(self,
                     times: List[float],
                     forces: List[float],
                     extensions: List[float],
                     stresses: List[float],
                     strains: List[float],
                     config: TestConfiguration,
                     properties: MechanicalProperties,
                     filepath: str = None) -> str:
        """Export test data to Excel file with formatting and charts."""
        if not HAS_OPENPYXL:
            raise ExportError("openpyxl not installed. Install with: pip install openpyxl")
        
        filename = filepath if filepath else self.generate_filename(config, "xlsx")
        
        try:
            wb = Workbook()
            
            # Styles
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            accent_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ===== Summary Sheet =====
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Title
            ws_summary['A1'] = "Tensile Test Report"
            ws_summary['A1'].font = title_font
            
            # Metadata section
            row = 3
            metadata_items = [
                ("Test Information", ""),
                ("Sample ID", config.metadata.sample_id),
                ("Material", config.metadata.material_name),
                ("Grade", config.metadata.material_grade),
                ("Standard", config.metadata.test_standard.value),
                ("Date", config.metadata.test_date),
                ("Time", config.metadata.test_time),
                ("Operator", config.metadata.operator_name),
                ("Customer", config.metadata.customer_name),
                ("", ""),
                ("Specimen Dimensions", ""),
                ("Gauge Length", f"{config.specimen.gauge_length} mm"),
                ("Thickness", f"{config.specimen.thickness} mm"),
                ("Width", f"{config.specimen.width} mm"),
                ("Cross-Section", f"{config.specimen.cross_section_area} mm²"),
                ("", ""),
                ("Environment", ""),
                ("Temperature", f"{config.metadata.temperature} °C"),
                ("Humidity", f"{config.metadata.humidity} % RH"),
            ]
            
            for label, value in metadata_items:
                if label and not value:  # Section header
                    ws_summary.cell(row=row, column=1, value=label).font = header_font
                else:
                    ws_summary.cell(row=row, column=1, value=label)
                    ws_summary.cell(row=row, column=2, value=value)
                row += 1
            
            # Results section
            row += 1
            ws_summary.cell(row=row, column=1, value="Mechanical Properties").font = header_font
            row += 1
            
            results_items = [
                ("Ultimate Tensile Strength", f"{properties.ultimate_tensile_strength:.2f}", "MPa"),
                ("Yield Strength (Rp0.2)", f"{properties.yield_strength_offset:.2f}", "MPa"),
                ("Young's Modulus", f"{properties.youngs_modulus:.1f}", "MPa"),
                ("Elongation at Break", f"{properties.strain_at_break:.2f}", "%"),
                ("Maximum Force", f"{properties.max_force:.2f}", "N"),
                ("Energy to Break", f"{properties.energy_to_break:.4f}", "J"),
            ]
            
            ws_summary.cell(row=row, column=1, value="Property").font = header_font
            ws_summary.cell(row=row, column=2, value="Value").font = header_font
            ws_summary.cell(row=row, column=3, value="Unit").font = header_font
            row += 1
            
            for prop, val, unit in results_items:
                ws_summary.cell(row=row, column=1, value=prop)
                ws_summary.cell(row=row, column=2, value=float(val.replace(',', '.')))
                ws_summary.cell(row=row, column=3, value=unit)
                row += 1
            
            # Adjust column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 10
            
            # ===== Data Sheet =====
            ws_data = wb.create_sheet("Raw Data")
            
            # Headers
            headers = ["Time (s)", "Force (N)", "Extension (mm)", "Stress (MPa)", "Strain (%)"]
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = accent_fill
            
            # Data
            for i in range(len(times)):
                ws_data.cell(row=i+2, column=1, value=times[i])
                ws_data.cell(row=i+2, column=2, value=forces[i])
                ws_data.cell(row=i+2, column=3, value=extensions[i])
                ws_data.cell(row=i+2, column=4, value=stresses[i])
                ws_data.cell(row=i+2, column=5, value=strains[i]*100)
            
            # Adjust column widths
            for col in range(1, 6):
                ws_data.column_dimensions[chr(64+col)].width = 15
            
            # ===== Chart Sheet =====
            ws_chart = wb.create_sheet("Charts")
            
            # Force vs Extension Chart
            chart1 = LineChart()
            chart1.title = "Force vs Extension"
            chart1.y_axis.title = "Force (N)"
            chart1.x_axis.title = "Extension (mm)"
            chart1.style = 10
            
            data = Reference(ws_data, min_col=2, min_row=1, max_row=len(times)+1)
            cats = Reference(ws_data, min_col=3, min_row=2, max_row=len(times)+1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.height = 15
            chart1.width = 20
            
            ws_chart.add_chart(chart1, "A1")
            
            # Stress vs Strain Chart
            chart2 = LineChart()
            chart2.title = "Stress vs Strain"
            chart2.y_axis.title = "Stress (MPa)"
            chart2.x_axis.title = "Strain (%)"
            chart2.style = 10
            
            data2 = Reference(ws_data, min_col=4, min_row=1, max_row=len(times)+1)
            cats2 = Reference(ws_data, min_col=5, min_row=2, max_row=len(times)+1)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            chart2.height = 15
            chart2.width = 20
            
            ws_chart.add_chart(chart2, "A32")
            
            wb.save(filename)
            return filename
            
        except Exception as e:
            raise ExportError(f"Excel export failed: {e}")
    
    # ============== JSON Export ==============
    
    def export_json(self,
                    times: List[float],
                    forces: List[float],
                    extensions: List[float],
                    stresses: List[float],
                    strains: List[float],
                    config: TestConfiguration,
                    properties: MechanicalProperties,
                    filepath: Optional[str] = None) -> str:
        """Export test data to JSON file."""
        filename = filepath if filepath else self.generate_filename(config, "json")
        
        try:
            # Build export data structure
            export_data = {
                "metadata": {
                    "test_id": config.metadata.test_id,
                    "sample_id": config.metadata.sample_id,
                    "batch_id": config.metadata.batch_id,
                    "material": {
                        "name": config.metadata.material_name,
                        "type": config.metadata.material_type.value,
                        "grade": config.metadata.material_grade,
                    },
                    "standard": config.metadata.test_standard.value,
                    "operator": config.metadata.operator_name,
                    "customer": config.metadata.customer_name,
                    "date": config.metadata.test_date,
                    "time": config.metadata.test_time,
                    "environment": {
                        "temperature_c": config.metadata.temperature,
                        "humidity_pct": config.metadata.humidity,
                    },
                    "notes": config.metadata.notes,
                },
                "specimen": {
                    "type": config.specimen.specimen_type,
                    "gauge_length_mm": config.specimen.gauge_length,
                    "thickness_mm": config.specimen.thickness,
                    "width_mm": config.specimen.width,
                    "cross_section_mm2": config.specimen.cross_section_area,
                    "grip_distance_mm": config.specimen.grip_distance,
                },
                "test_parameters": {
                    "control_mode": config.control.control_mode.value,
                    "test_speed_mm_min": config.control.test_speed,
                    "preload_n": config.control.preload_value if config.control.preload_enabled else 0,
                },
                "results": {
                    "strength": {
                        "ultimate_tensile_strength_mpa": round(properties.ultimate_tensile_strength, 2),
                        "yield_strength_mpa": round(properties.yield_strength_offset, 2),
                        "break_stress_mpa": round(properties.break_stress, 2),
                    },
                    "force": {
                        "max_force_n": round(properties.max_force, 2),
                        "force_at_yield_n": round(properties.force_at_yield, 2),
                        "force_at_break_n": round(properties.force_at_break, 2),
                    },
                    "modulus": {
                        "youngs_modulus_mpa": round(properties.youngs_modulus, 1),
                        "r_squared": round(properties.modulus_r_squared, 4),
                    },
                    "strain": {
                        "elongation_at_break_pct": round(properties.strain_at_break, 2),
                        "strain_at_yield_pct": round(properties.strain_at_yield, 2),
                        "strain_at_uts_pct": round(properties.strain_at_uts, 2),
                    },
                    "energy": {
                        "energy_to_break_j": round(properties.energy_to_break, 4),
                        "energy_to_yield_j": round(properties.energy_to_yield, 4),
                    },
                },
                "raw_data": {
                    "points": len(times),
                    "columns": ["time_s", "force_n", "extension_mm", "stress_mpa", "strain_pct"],
                    "data": [
                        [round(t, 4), round(f, 4), round(e, 4), round(s, 4), round(st*100, 4)]
                        for t, f, e, s, st in zip(times, forces, extensions, stresses, strains)
                    ]
                },
                "export_info": {
                    "format_version": "1.0",
                    "exported_at": datetime.now().isoformat(),
                    "software": "DIY Tensile Tester v2.0",
                }
            }
            
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)
            
            return filename
            
        except Exception as e:
            raise ExportError(f"JSON export failed: {e}")
    
    # ============== PDF Export ==============
    
    def export_pdf(self,
                   times: List[float],
                   forces: List[float],
                   extensions: List[float],
                   stresses: List[float],
                   strains: List[float],
                   config: TestConfiguration,
                   properties: MechanicalProperties,
                   include_plot: bool = True,
                   filepath: Optional[str] = None) -> str:
        """Export test report to PDF file."""
        if not HAS_REPORTLAB:
            raise ExportError("reportlab not installed. Install with: pip install reportlab")
        
        filename = filepath if filepath else self.generate_filename(config, "pdf")
        
        try:
            doc = SimpleDocTemplate(
                filename,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=20*mm,
                bottomMargin=20*mm
            )
            
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=20,
                alignment=1  # Center
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=12,
                spaceBefore=15,
                spaceAfter=10,
                textColor=colors.HexColor('#1976D2')
            )
            
            normal_style = styles['Normal']
            
            # Build content
            content = []
            
            # Title
            content.append(Paragraph("Tensile Test Report", title_style))
            content.append(Spacer(1, 10))
            
            # Test Information
            content.append(Paragraph("Test Information", heading_style))
            
            info_data = [
                ["Sample ID:", config.metadata.sample_id or "N/A"],
                ["Material:", config.metadata.material_name or "N/A"],
                ["Grade:", config.metadata.material_grade or "N/A"],
                ["Standard:", config.metadata.test_standard.value],
                ["Date:", config.metadata.test_date],
                ["Operator:", config.metadata.operator_name or "N/A"],
                ["Customer:", config.metadata.customer_name or "N/A"],
            ]
            
            info_table = Table(info_data, colWidths=[120, 300])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            content.append(info_table)
            
            # Specimen Dimensions
            content.append(Paragraph("Specimen Dimensions", heading_style))
            
            spec_data = [
                ["Specimen Type:", config.specimen.specimen_type],
                ["Gauge Length:", f"{config.specimen.gauge_length} mm"],
                ["Thickness:", f"{config.specimen.thickness} mm"],
                ["Width:", f"{config.specimen.width} mm"],
                ["Cross-Section Area:", f"{config.specimen.cross_section_area} mm²"],
                ["Grip Distance:", f"{config.specimen.grip_distance} mm"],
            ]
            
            spec_table = Table(spec_data, colWidths=[120, 300])
            spec_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            content.append(spec_table)
            
            # Mechanical Properties
            content.append(Paragraph("Mechanical Properties", heading_style))
            
            # Results table
            results_data = [
                ["Property", "Symbol", "Value", "Unit"],
                ["Ultimate Tensile Strength", "σ_UTS", f"{properties.ultimate_tensile_strength:.2f}", "MPa"],
                ["Yield Strength (Rp0.2)", "σ_y", f"{properties.yield_strength_offset:.2f}", "MPa"],
                ["Young's Modulus", "E", f"{properties.youngs_modulus:.1f}", "MPa"],
                ["Elongation at Break", "ε_b", f"{properties.strain_at_break:.2f}", "%"],
                ["Maximum Force", "F_max", f"{properties.max_force:.2f}", "N"],
                ["Force at Yield", "F_y", f"{properties.force_at_yield:.2f}", "N"],
                ["Energy to Break", "U_b", f"{properties.energy_to_break:.4f}", "J"],
            ]
            
            results_table = Table(results_data, colWidths=[150, 50, 80, 50])
            results_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            content.append(results_table)
            
            # Test Conditions
            content.append(Paragraph("Test Conditions", heading_style))
            
            cond_data = [
                ["Control Mode:", config.control.control_mode.value],
                ["Test Speed:", f"{config.control.test_speed} mm/min"],
                ["Temperature:", f"{config.metadata.temperature} °C"],
                ["Humidity:", f"{config.metadata.humidity} % RH"],
                ["Data Points:", str(len(times))],
                ["Test Duration:", f"{times[-1] if times else 0:.1f} s"],
            ]
            
            cond_table = Table(cond_data, colWidths=[120, 300])
            cond_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            content.append(cond_table)
            
            # Notes
            if config.metadata.notes:
                content.append(Paragraph("Notes", heading_style))
                content.append(Paragraph(config.metadata.notes, normal_style))
            
            # Signature fields
            content.append(Spacer(1, 40))
            sig_data = [
                ["", ""],
                ["_" * 30, "_" * 30],
                ["Tested by", "Approved by"],
            ]
            sig_table = Table(sig_data, colWidths=[200, 200])
            sig_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 2), (-1, 2), 5),
            ]))
            content.append(sig_table)
            
            # Footer info
            content.append(Spacer(1, 20))
            footer = Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
                f"Software: DIY Tensile Tester v2.0",
                ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.grey)
            )
            content.append(footer)
            
            doc.build(content)
            return filename
            
        except Exception as e:
            raise ExportError(f"PDF export failed: {e}")
    
    # ============== XML Export ==============
    
    def export_xml(self,
                   times: List[float],
                   forces: List[float],
                   extensions: List[float],
                   stresses: List[float],
                   strains: List[float],
                   config: TestConfiguration,
                   properties: MechanicalProperties) -> str:
        """Export test data to XML file."""
        filename = self.generate_filename(config, "xml")
        
        try:
            # Build XML manually for simplicity
            xml_lines = [
                '<?xml version="1.0" encoding="UTF-8"?>',
                '<TensileTestReport>',
                '  <Metadata>',
                f'    <TestID>{config.metadata.test_id}</TestID>',
                f'    <SampleID>{config.metadata.sample_id}</SampleID>',
                f'    <BatchID>{config.metadata.batch_id}</BatchID>',
                f'    <MaterialName>{config.metadata.material_name}</MaterialName>',
                f'    <MaterialType>{config.metadata.material_type.value}</MaterialType>',
                f'    <Standard>{config.metadata.test_standard.value}</Standard>',
                f'    <Operator>{config.metadata.operator_name}</Operator>',
                f'    <Customer>{config.metadata.customer_name}</Customer>',
                f'    <TestDate>{config.metadata.test_date}</TestDate>',
                f'    <TestTime>{config.metadata.test_time}</TestTime>',
                f'    <Temperature unit="C">{config.metadata.temperature}</Temperature>',
                f'    <Humidity unit="percent">{config.metadata.humidity}</Humidity>',
                '  </Metadata>',
                '  <Specimen>',
                f'    <Type>{config.specimen.specimen_type}</Type>',
                f'    <GaugeLength unit="mm">{config.specimen.gauge_length}</GaugeLength>',
                f'    <Thickness unit="mm">{config.specimen.thickness}</Thickness>',
                f'    <Width unit="mm">{config.specimen.width}</Width>',
                f'    <CrossSection unit="mm2">{config.specimen.cross_section_area}</CrossSection>',
                '  </Specimen>',
                '  <Results>',
                '    <Strength>',
                f'      <UltimateTensileStrength unit="MPa">{properties.ultimate_tensile_strength:.2f}</UltimateTensileStrength>',
                f'      <YieldStrength unit="MPa">{properties.yield_strength_offset:.2f}</YieldStrength>',
                f'      <BreakStress unit="MPa">{properties.break_stress:.2f}</BreakStress>',
                '    </Strength>',
                '    <Force>',
                f'      <MaxForce unit="N">{properties.max_force:.2f}</MaxForce>',
                f'      <ForceAtYield unit="N">{properties.force_at_yield:.2f}</ForceAtYield>',
                f'      <ForceAtBreak unit="N">{properties.force_at_break:.2f}</ForceAtBreak>',
                '    </Force>',
                '    <Modulus>',
                f'      <YoungsModulus unit="MPa">{properties.youngs_modulus:.1f}</YoungsModulus>',
                f'      <RSquared>{properties.modulus_r_squared:.4f}</RSquared>',
                '    </Modulus>',
                '    <Strain>',
                f'      <ElongationAtBreak unit="percent">{properties.strain_at_break:.2f}</ElongationAtBreak>',
                f'      <StrainAtYield unit="percent">{properties.strain_at_yield:.2f}</StrainAtYield>',
                '    </Strain>',
                '    <Energy>',
                f'      <EnergyToBreak unit="J">{properties.energy_to_break:.4f}</EnergyToBreak>',
                f'      <EnergyToYield unit="J">{properties.energy_to_yield:.4f}</EnergyToYield>',
                '    </Energy>',
                '  </Results>',
                '  <RawData>',
                f'    <Points>{len(times)}</Points>',
                '    <DataPoints>',
            ]
            
            # Add data points
            for i in range(len(times)):
                xml_lines.append(
                    f'      <Point time="{times[i]:.4f}" force="{forces[i]:.4f}" '
                    f'extension="{extensions[i]:.4f}" stress="{stresses[i]:.4f}" '
                    f'strain="{strains[i]*100:.4f}"/>'
                )
            
            xml_lines.extend([
                '    </DataPoints>',
                '  </RawData>',
                '  <ExportInfo>',
                f'    <ExportedAt>{datetime.now().isoformat()}</ExportedAt>',
                '    <Software>DIY Tensile Tester v2.0</Software>',
                '  </ExportInfo>',
                '</TensileTestReport>',
            ])
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write('\n'.join(xml_lines))
            
            return filename
            
        except Exception as e:
            raise ExportError(f"XML export failed: {e}")
    
    # ============== Batch Export ==============
    
    def export_all(self,
                   times: List[float],
                   forces: List[float],
                   extensions: List[float],
                   stresses: List[float],
                   strains: List[float],
                   config: TestConfiguration,
                   properties: MechanicalProperties,
                   export_config: ExportConfig) -> Dict[str, str]:
        """Export to all selected formats."""
        results = {}
        
        if export_config.export_csv:
            try:
                results['csv'] = self.export_csv(
                    times, forces, extensions, stresses, strains,
                    config, properties
                )
            except ExportError as e:
                results['csv_error'] = str(e)
        
        if export_config.export_excel:
            try:
                results['excel'] = self.export_excel(
                    times, forces, extensions, stresses, strains,
                    config, properties
                )
            except ExportError as e:
                results['excel_error'] = str(e)
        
        if export_config.export_json:
            try:
                results['json'] = self.export_json(
                    times, forces, extensions, stresses, strains,
                    config, properties
                )
            except ExportError as e:
                results['json_error'] = str(e)
        
        if export_config.export_pdf:
            try:
                results['pdf'] = self.export_pdf(
                    times, forces, extensions, stresses, strains,
                    config, properties
                )
            except ExportError as e:
                results['pdf_error'] = str(e)
        
        if export_config.export_xml:
            try:
                results['xml'] = self.export_xml(
                    times, forces, extensions, stresses, strains,
                    config, properties
                )
            except ExportError as e:
                results['xml_error'] = str(e)
        
        return results
